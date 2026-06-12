import os
import sqlite3
import openpyxl

excel_path = "/home/guss/Descargas/TRAMITE/BLOQUE I/POLLA/Polla_Mundialista_Consolidada.xlsx"
db_path = "/home/guss/Documentos/POLLA/APP/db/database.db"

# Ensure directories exist
os.makedirs(os.path.dirname(db_path), exist_ok=True)

print("Connecting to database...")
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Create tables
cursor.execute("DROP TABLE IF EXISTS predictions")
cursor.execute("DROP TABLE IF EXISTS matches")
cursor.execute("DROP TABLE IF EXISTS participants")

cursor.execute("""
CREATE TABLE participants (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
)
""")

cursor.execute("""
CREATE TABLE matches (
    id INTEGER PRIMARY KEY,
    group_name TEXT NOT NULL,
    home_team TEXT NOT NULL,
    away_team TEXT NOT NULL,
    home_actual INTEGER,
    away_actual INTEGER
)
""")

cursor.execute("""
CREATE TABLE predictions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    participant_id INTEGER NOT NULL,
    match_id INTEGER NOT NULL,
    home_pred INTEGER,
    away_pred INTEGER,
    FOREIGN KEY(participant_id) REFERENCES participants(id),
    FOREIGN KEY(match_id) REFERENCES matches(id)
)
""")

print("Loading Excel workbook...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb["Consolidado"]

# 1. Parse and insert matches
print("Importing matches...")
matches_list = []
for r in range(6, 78):
    m_id = ws.cell(row=r, column=1).value
    group = ws.cell(row=r, column=2).value
    home = ws.cell(row=r, column=3).value
    away = ws.cell(row=r, column=4).value
    
    if m_id is not None:
        m_id = int(m_id)
        cursor.execute(
            "INSERT INTO matches (id, group_name, home_team, away_team, home_actual, away_actual) VALUES (?, ?, ?, ?, NULL, NULL)",
            (m_id, str(group), str(home), str(away))
        )
        matches_list.append(m_id)

# 2. Parse and insert participants
print("Importing participants...")
participant_cols = {}
max_col = ws.max_column
for c in range(5, max_col + 1, 2):
    p_name = ws.cell(row=4, column=c).value
    if p_name:
        cursor.execute("INSERT OR IGNORE INTO participants (name) VALUES (?)", (str(p_name),))
        conn.commit()
        
        # Get participant id
        cursor.execute("SELECT id FROM participants WHERE name = ?", (str(p_name),))
        p_id = cursor.fetchone()[0]
        
        participant_cols[p_name] = (p_id, c)
        print(f"Participant: {p_name} (ID: {p_id}, Col: {c})")

# 3. Parse and insert predictions
print("Importing predictions...")
for p_name, (p_id, col_idx) in participant_cols.items():
    for r in range(6, 78):
        m_id = int(ws.cell(row=r, column=1).value)
        home_pred = ws.cell(row=r, column=col_idx).value
        away_pred = ws.cell(row=r, column=col_idx+1).value
        
        # Convert to int if not None
        if home_pred is not None:
            try:
                home_pred = int(home_pred)
            except ValueError:
                home_pred = None
        if away_pred is not None:
            try:
                away_pred = int(away_pred)
            except ValueError:
                away_pred = None
                
        cursor.execute(
            "INSERT INTO predictions (participant_id, match_id, home_pred, away_pred) VALUES (?, ?, ?, ?)",
            (p_id, m_id, home_pred, away_pred)
        )

conn.commit()
conn.close()
print("Database seeding completed successfully!")
